from openpyxl import load_workbook
from pathlib import Path
from .data_models import ExcelSheetData, Transaction, BaibitTransaction
import logging

logger = logging.getLogger(__name__)


class ExcelProcessor:
    @staticmethod
    def process_workbook(file_path: Path, agent_percent: float) -> dict[str, ExcelSheetData]:
        """
        Обрабатывает Excel-файл и возвращает данные по каждому листу

        Args:
            file_path: Путь к файлу Excel
            agent_percent: Процент агента для расчетов

        Returns:
            Словарь {название_листа: данные_листа}
        """
        try:
            wb = load_workbook(file_path, data_only=True)
            sheets_data = {}

            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]

                # Основные данные из фиксированных ячеек
                data = ExcelSheetData(
                    sheet_name=sheet_name,
                    full_name=sheet['K2'].value or "Не указано",
                    bank=sheet['L2'].value or "Не указан",
                    warm_up_purchases=sheet['M2'].value or 0,
                    warm_up_rub=sheet['N2'].value or 0,
                    start_balance=sheet['Q2'].value or 0,
                    stop_balance=sheet['R2'].value or 0,
                    start_time=str(sheet['S2'].value) if sheet['S2'].value else "Нет данных",
                    end_time=str(sheet['T2'].value) if sheet['T2'].value else "Нет данных",
                    operator=sheet['P2'].value or "Не указан",
                    inflows=[],
                    outflows=[],
                    baibit=[],
                    agent_percent=agent_percent
                )

                # Обработка строк с транзакциями
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    try:
                        # Входящие транзакции (столбцы A и B)
                        if row[0] and row[1]:
                            data.inflows.append(Transaction(
                                amount=float(row[0]),
                                transaction_id=str(row[1])
                            ))
                            data.turnover += float(row[0])

                        # Исходящие транзакции (столбцы C, D и E - комиссия)
                        if row[2] and row[3]:
                            commission = float(row[4]) if len(row) > 4 and row[4] else 0
                            data.outflows.append(Transaction(
                                amount=float(row[2]),
                                transaction_id=str(row[3]),
                                commission=commission
                            ))

                        # Байбит транзакции (столбцы F и G)
                        if row[5] and row[6]:
                            data.baibit.append(BaibitTransaction(
                                amount=float(row[5]),
                                rate=float(row[6])
                            ))

                    except (ValueError, TypeError) as e:
                        logger.warning(f"Ошибка обработки строки: {row}. {str(e)}")
                        continue

                # Расчет выплат
                data.calculate_payments()
                sheets_data[sheet_name] = data

            return sheets_data

        except Exception as e:
            logger.error(f"Ошибка обработки файла {file_path}: {str(e)}", exc_info=True)
            raise ValueError(f"Невозможно обработать файл: {str(e)}")